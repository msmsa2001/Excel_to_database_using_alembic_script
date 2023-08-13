"""Script Data Upload

Revision ID: 401fb4340d14
Revises: 5c6af7a17936
Create Date: 2023-08-13 12:13:28.312344

"""
from typing import Sequence, Union
import openpyxl

from run import Topic, SubTopic
from sqlalchemy.orm import Session

from alembic import op
import sqlalchemy as sa


# revision identifiers, used by Alembic.
revision: str = '401fb4340d14'
down_revision: Union[str, None] = '5c6af7a17936'
branch_labels: Union[str, Sequence[str], None] = None
depends_on: Union[str, Sequence[str], None] = None


def upgrade() -> None:
    topic=[]
    subtopic=[]

    workbook=openpyxl.load_workbook('Data.xlsx')
    sheet=workbook.sheetnames
    current_sheet_name=workbook[sheet[0]]

    topic=[{"id":index, "name":i.value} for index, i in enumerate(current_sheet_name[1],start=1)]

    subtopic_id=0
    topic_id=1
    for col in current_sheet_name.iter_cols(values_only=True):
        new_col = [i for i in col[1:] if i is not None]
        c=[{"id":index+subtopic_id, "name":i,"topic_id":topic_id} for index,i in enumerate(new_col,start=1)]
        subtopic_id += len(c)
        topic_id += 1
        subtopic.extend(c)

        
    conn = op.get_bind()
    session = Session(bind=conn)
    session.bulk_insert_mappings(SubTopic, subtopic)
    session.bulk_insert_mappings(Topic, topic)

    session.commit()


def downgrade() -> None:
    conn = op.get_bind()

    conn.execute(
        sa.text(
            """
            TRUNCATE TABLE topic_name cascade;
            TRUNCATE TABLE subtopic_name cascade;
            """
        )
    )
